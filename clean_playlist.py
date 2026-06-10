import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ─────────────────────────────────────────────
# 1. DATA INGESTION & VALIDATION
# ─────────────────────────────────────────────

df = pd.read_csv('Atlantic_United_States.csv')

validation_log = []

# 1a. Parse date (DD-MM-YYYY)
df['date'] = pd.to_datetime(df['date'], format='%d-%m-%Y', errors='coerce')
invalid_dates = df['date'].isna().sum()
validation_log.append(("Invalid date values", invalid_dates))

# 1b. Validate rank range 1–50
out_of_range = ((df['position'] < 1) | (df['position'] > 50)).sum()
validation_log.append(("Positions outside 1–50", out_of_range))

# 1c. Identify duplicate song-date entries
# Note: 01-03-2025 has 100 rows (full playlist duplicated in source)
# For same song+artist on same date keep first occurrence
before_dedup = len(df)
df = df.sort_values(['date', 'position'])
df = df.drop_duplicates(subset=['song', 'artist', 'date'], keep='first')
after_dedup = len(df)
validation_log.append(("Duplicate song-date rows removed", before_dedup - after_dedup))

# Ghostbusters & Christmas songs are different artists – not truly duplicate, keep both

# 1d. Check missing values
missing = df.isnull().sum().sum()
validation_log.append(("Missing values after cleaning", missing))

# 1e. Standardize artist name formatting
# Preserve known stylistic all-caps (SZA, BTS, KAROL G, JVKE etc.) and all-lower
# Rule: strip leading/trailing whitespace, normalize multiple spaces
df['artist'] = df['artist'].str.strip().str.replace(r'\s+', ' ', regex=True)
df['song'] = df['song'].str.strip().str.replace(r'\s+', ' ', regex=True)

# ─────────────────────────────────────────────
# 2. FEATURE ENGINEERING
# ─────────────────────────────────────────────

# Sort for rolling calculations
df = df.sort_values(['song', 'artist', 'date']).reset_index(drop=True)

# 2a. Duration in minutes
df['duration_min'] = (df['duration_ms'] / 60000).round(2)

# 2b. Group-level metrics per unique song+artist
song_groups = df.groupby(['song', 'artist'])

agg = song_groups.agg(
    days_on_chart=('date', 'nunique'),
    avg_rank=('position', 'mean'),
    best_rank=('position', 'min'),
    rank_volatility_index=('position', 'std'),
).round(2).reset_index()

agg['avg_rank'] = agg['avg_rank'].round(2)
agg['rank_volatility_index'] = agg['rank_volatility_index'].fillna(0).round(2)

# 2c. Popularity Trend Score – 7-day rolling average of popularity
df = df.sort_values(['song', 'artist', 'date']).reset_index(drop=True)
df['popularity_trend_score'] = (
    df.groupby(['song', 'artist'])['popularity']
    .transform(lambda x: x.rolling(window=7, min_periods=1).mean())
    .round(2)
)

# 2d. Merge aggregated metrics back into main df
df = df.merge(agg, on=['song', 'artist'], how='left')

# Final column order
col_order = [
    'date', 'position', 'song', 'artist',
    'popularity', 'popularity_trend_score',
    'duration_ms', 'duration_min',
    'days_on_chart', 'avg_rank', 'best_rank', 'rank_volatility_index',
    'album_type', 'total_tracks', 'is_explicit', 'album_cover_url'
]
df = df[col_order].sort_values(['date', 'position']).reset_index(drop=True)

# ─────────────────────────────────────────────
# 3. WRITE EXCEL OUTPUT
# ─────────────────────────────────────────────

wb = Workbook()

# ── SHEET 1: Cleaned Data ──────────────────
ws1 = wb.active
ws1.title = "Cleaned Data"

# Styles
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
header_fill = PatternFill('solid', start_color='1F4E79')
alt_fill = PatternFill('solid', start_color='D6E4F0')
normal_fill = PatternFill('solid', start_color='FFFFFF')
center = Alignment(horizontal='center', vertical='center')
left = Alignment(horizontal='left', vertical='center')
thin = Side(style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = [
    'Date', 'Position', 'Song', 'Artist',
    'Popularity', 'Popularity Trend Score (7d)',
    'Duration (ms)', 'Duration (min)',
    'Days on Chart', 'Avg Rank', 'Best Rank', 'Rank Volatility Index',
    'Album Type', 'Total Tracks', 'Is Explicit', 'Album Cover URL'
]

for col_idx, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

for r_idx, row in enumerate(df.itertuples(index=False), 2):
    fill = alt_fill if r_idx % 2 == 0 else normal_fill
    for c_idx, val in enumerate(row, 1):
        if isinstance(val, pd.Timestamp):
            val = val.strftime('%d-%m-%Y')
        cell = ws1.cell(row=r_idx, column=c_idx, value=val)
        cell.font = Font(name='Arial', size=9)
        cell.fill = fill
        cell.border = border
        cell.alignment = center if c_idx != 3 else left

col_widths = [13, 10, 38, 30, 12, 26, 14, 14, 14, 10, 10, 22, 12, 13, 12, 50]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.freeze_panes = 'A2'
ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

# ── SHEET 2: Song Summary ──────────────────
ws2 = wb.create_sheet("Song Summary")

summary_df = df.groupby(['song', 'artist']).agg(
    days_on_chart=('days_on_chart', 'first'),
    avg_rank=('avg_rank', 'first'),
    best_rank=('best_rank', 'first'),
    rank_volatility_index=('rank_volatility_index', 'first'),
    avg_popularity=('popularity', 'mean'),
    duration_min=('duration_min', 'first'),
    album_type=('album_type', 'first'),
    is_explicit=('is_explicit', 'first'),
).round(2).reset_index().sort_values('days_on_chart', ascending=False)

sum_headers = [
    'Song', 'Artist', 'Days on Chart', 'Avg Rank', 'Best Rank',
    'Rank Volatility Index', 'Avg Popularity', 'Duration (min)',
    'Album Type', 'Is Explicit'
]

for col_idx, header in enumerate(sum_headers, 1):
    cell = ws2.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = PatternFill('solid', start_color='375623')
    cell.alignment = center
    cell.border = border

for r_idx, row in enumerate(summary_df.itertuples(index=False), 2):
    fill = PatternFill('solid', start_color='E2EFDA') if r_idx % 2 == 0 else normal_fill
    for c_idx, val in enumerate(row, 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=val)
        cell.font = Font(name='Arial', size=9)
        cell.fill = fill
        cell.border = border
        cell.alignment = center if c_idx != 1 else left

sum_widths = [38, 30, 14, 10, 10, 22, 15, 14, 12, 12]
for i, w in enumerate(sum_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = f"A1:{get_column_letter(len(sum_headers))}1"

# ── SHEET 3: Validation Log ───────────────
ws3 = wb.create_sheet("Validation Log")

ws3['A1'] = 'Data Ingestion & Validation Report'
ws3['A1'].font = Font(name='Arial', bold=True, size=13, color='1F4E79')
ws3.merge_cells('A1:C1')

vlog_headers = ['Check', 'Count / Result', 'Status']
for c, h in enumerate(vlog_headers, 1):
    cell = ws3.cell(row=3, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

checks = validation_log + [
    ("Total rows after cleaning", len(df)),
    ("Unique songs", df['song'].nunique()),
    ("Unique artists", df['artist'].nunique()),
    ("Date range start", df['date'].min().strftime('%d-%m-%Y')),
    ("Date range end", df['date'].max().strftime('%d-%m-%Y')),
    ("Unique chart dates", df['date'].nunique()),
]

for r, (check, result) in enumerate(checks, 4):
    if isinstance(result, int):
        status = '✅ OK' if result == 0 or check.startswith('Total') or check.startswith('Unique') or check == 'Unique chart dates' else '⚠️ Handled'
    else:
        status = '✅ OK'
    for c, val in enumerate([check, result, status], 1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.font = Font(name='Arial', size=10)
        cell.border = border
        cell.alignment = center if c != 1 else left
        if status == '✅ OK':
            cell.fill = PatternFill('solid', start_color='E2EFDA')
        else:
            cell.fill = PatternFill('solid', start_color='FFF2CC')

for col, w in zip(['A', 'B', 'C'], [38, 20, 15]):
    ws3.column_dimensions[col].width = w

wb.save('/home/claude/US_Top50_Cleaned.xlsx')
print("✅ Excel file saved successfully.")
print(f"   Rows in cleaned data: {len(df)}")
print(f"   Unique songs tracked: {df['song'].nunique()}")
print(f"   Feature columns added: days_on_chart, avg_rank, best_rank, rank_volatility_index, popularity_trend_score, duration_min")
